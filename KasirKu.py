from tkinter import *
from openpyxl import workbook, load_workbook

# globally declare the expression variable
expression = ""
harga = ""

def intoxl():
	T=expression_field.get()
	G=nam_pembeli.get()
	U=tanggal.get()
	wb = load_workbook('yourfile.xlsx')
	ws = wb.active
	ws.cell(column=1,row=ws.max_row+1, value=T)
	ws.cell(column=2,row=ws.max_row, value=G)
	ws.cell(column=3,row=ws.max_row, value=U)

	wb.save('yourfile.xlsx')


# Function to update expressiom
# in the text entry box
def press(num):
	# point out the global expression variable
	global expression
	global harga
	# concatenation of string

	expression = expression + str(num)
	if num == 1:
		harga = harga + str(30.000)
	if num == 2:
		harga = harga + str(28.000)
	if num == 3:
		harga = harga + str(18.000)
	if num == 4:
		harga = harga + str(25.000)
	if num == 5:
		harga = harga + str(25.000)
	if num == 6:
		harga = harga + str(15.000)
	if num == 7:
		harga = harga + str(18.000)
	if num == 8:
		harga = harga + str(20.000)
	if num == 9:
		harga = harga + str(12.000)
	if num == 10:
		harga = harga + str(15.000)
	if num == 11:
		harga = harga + str(12.000)
	if num == 12:
		harga = harga + str(17.000)
	if num == '+':
		harga = harga + str(num)

	# update the expression by using set method
	equation.set(expression)


# Function to evaluate the final expression
def equalpress():
	# Try and except statement is used
	# for handling the errors like zero
	# division error etc.

	# Put that code inside the try block
	# which may generate the error
	try:

		global expression
		global harga

		# eval function evaluate the expression
		# and str function convert the result
		# into string
		total = str(eval(harga))

		equation.set(total)

		

		# initialze the expression variable
		# by empty string
		expression = ""
		harga = ""

	# if error is generate then handle
	# by the except block
	except:

		equation.set(" error babi ")
		expression = ""


	
# Function to clear the contents
# of text entry box
def clear():
	global expression
	expression = ""
	equation.set("")
	global harga
	harga = ""
	harga.set("")


# Driver code
if __name__ == "__main__":
	# create a GUI window
	gui = Tk()

	# set the background colour of GUI window
	gui.configure(background="dark blue")

	# set the title of GUI window
	gui.title("Kasirku")

	# set the configuration of GUI window
	gui.geometry("1100x800")

	# StringVar() is the variable class
	# we create an instance of this class
	equation = StringVar()

	# create the text entry box for
	# showing the expression .
	expression_field = Entry(gui,width = 20, textvariable=equation)
	nam_pembeli = Entry(gui,width = 10)
	tanggal= Entry(gui,width = 10)
	
	L4 = Label(gui, text='<<< TANGGAL PEMBELIAN').grid(column=201,row=1)
	L5 = Label(gui, text='<<< NAMA PEMBELI').grid(column=100,row=1)


	

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	expression_field.grid(columnspan=4, ipadx=70,row=1)
	nam_pembeli.grid(column=40, ipadx=70,row=1)
	tanggal.grid(column=200, ipadx=70,row=1)

	# create a Buttons and place at a particular
	# location inside the root window .
	# when user press the button, the command or
	# function affiliated to that button is executed .
	button1 = Button(gui, text=' 1 ', fg='black', bg='red',
					command=lambda: press(1), height=1, width=7)
	button1.grid(row=2, column=0)

	button2 = Button(gui, text=' 2 ', fg='black', bg='red',
					command=lambda: press(2), height=1, width=7)
	button2.grid(row=2, column=1)

	button3 = Button(gui, text=' 3 ', fg='black', bg='red',
					command=lambda: press(3), height=1, width=7)
	button3.grid(row=2, column=2)

	button4 = Button(gui, text=' 4 ', fg='black', bg='red',
					command=lambda: press(4), height=1, width=7)
	button4.grid(row=3, column=0)

	button5 = Button(gui, text=' 5 ', fg='black', bg='red',
					command=lambda: press(5), height=1, width=7)
	button5.grid(row=3, column=1)

	button6 = Button(gui, text=' 6 ', fg='black', bg='red',
					command=lambda: press(6), height=1, width=7)
	button6.grid(row=3, column=2)

	button7 = Button(gui, text=' 7 ', fg='black', bg='red',
					command=lambda: press(7), height=1, width=7)
	button7.grid(row=4, column=0)

	button8 = Button(gui, text=' 8 ', fg='black', bg='red',
					command=lambda: press(8), height=1, width=7)
	button8.grid(row=4, column=1)

	button9 = Button(gui, text=' 9 ', fg='black', bg='red',
					command=lambda: press(9), height=1, width=7)
	button9.grid(row=4, column=2)

	button0 = Button(gui, text=' 0 ', fg='black', bg='red',
					command=lambda: press(0), height=1, width=7)
	button0.grid(row=5, column=0)

	plus = Button(gui, text=' + ', fg='black', bg='red',
				command=lambda: press("+"), height=1, width=7)
	plus.grid(row=2, column=3)

	minus = Button(gui, text=' - ', fg='black', bg='red',
				command=lambda: press("-"), height=1, width=7)
	minus.grid(row=3, column=3)

	multiply = Button(gui, text=' x ', fg='black', bg='red',
					command=lambda: press("*"), height=1, width=7)
	multiply.grid(row=4, column=3)

	divide = Button(gui, text=' / ', fg='black', bg='red',
					command=lambda: press("/"), height=1, width=7)
	divide.grid(row=5, column=3)

	equal = Button(gui, text=' = ', fg='black', bg='red',
				command=equalpress, height=1, width=7)
	equal.grid(row=5, column=2)

	input_toxl = Button(gui, text='save', command=intoxl)
	input_toxl.grid(row=5, column=3)

	clear = Button(gui, text='Clear', fg='black', bg='red',
				command=clear, height=1, width=7)
	clear.grid(row=5, column='1')

	Decimal= Button(gui, text='.', fg='black', bg='red',
					command=lambda: press('.'), height=1, width=7)
	Decimal.grid(row=6, column=0)

	# start the GUI
	gui.mainloop()
